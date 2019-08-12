# 셀 하나씩 가져오기
from openpyxl import load_workbook

wb = load_workbook('ex.xlsx')
data = wb.active

print(data['A1'].value)
print(data['A2'].value)
print(data['B1'].value)
print(data['B2'].value)

# 여러개의 셀 한꺼번에 가져오기
from openpyxl import load_workbook

wb = load_workbook('ex.xlsx')
data = wb.active

row = data['2']
for cell in row:
    print(cell.value)

print("-"*20)

col = data['A']
for cell in col:
    print(cell.value)

# 여러 행, 여러 열, 일정 영역, 다른 시트도 가져오기
from openpyxl import load_workbook

wb = load_workbook('ex.xlsx')

data = wb['test']

area = data['A1:B2']
for row in area:
    for cell in row:
        print(cell.value)

print('-'*20)

cols = data['A:B']
for col in cols:
    for cell in col:
        print(cell.value)

print('-'*20)

rows = data['1:2']
for row in rows:
    for cell in row:
        print(cell.value)