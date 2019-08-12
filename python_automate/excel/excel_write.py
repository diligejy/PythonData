from openpyxl import Workbook

wb = Workbook() # 클래스 변수 생성
ws = wb.create_sheet('sheet_test2')

ws['A1'] = 'jinyoung'
ws['B1'] = 'test'

wb.save('result.xlsx')